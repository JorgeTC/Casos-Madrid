import concurrent.futures
import datetime
import os
from bs4 import BeautifulSoup

import pdfplumber
import requests
from openpyxl import load_workbook


class URLFormat():

    def __init__(self, prefix, sufix, extension, date_format):
        self.prefix = prefix
        self.sufix = sufix
        self.extension = extension
        self.date_format = date_format

    def get_url(self, date):
        return self.prefix + self.date_format(date) + self.sufix + self.extension


class Downloader():

    SZ_CAM_URL = 'https://www.comunidad.madrid'
    SZ_ACTUAL_SITUATION = SZ_CAM_URL + '/servicios/salud/coronavirus#datos-situacion-actual'
    SZ_CAM_FILES = SZ_CAM_URL + "/sites/default/files"

    def __init__(self, date=datetime.date.today()):
        # Fecha que me interesa leer
        self.date = date

        # Nombre con el que guardo el PDF temporal
        self.pdf_name = 'tmp.pdf'

    def download_pdf(self):
        # Intento obtener el link desde la página de la CAM
        response = self.__get_pdf_response_fromCAM()

        # Si no lo he conseguido, pruebo las combinaciones de links
        if not response:
            response = self.__try_combinations()

        with open(self.pdf_name, 'wb') as f:
            # Descargo el PDF
            f.write(response.content)

    def __get_pdf_response_fromCAM(self):
        # Accedo a la página de la CAM donde está el enlace a la situación actual
        response = requests.get(self.SZ_ACTUAL_SITUATION)
        parsed = BeautifulSoup(response.text, 'html.parser')
        # Busco los div de la misma clase que contiene el link
        div_list = parsed.find_all('div', {'class': "field-item even"})

        # Itero todos los elementos que ha encontrado.
        # Sólo uno de ellos es válido
        for div in div_list:
            try:
                # Compruebo que sea un h3
                first_element = div.contents[0]
                if first_element.name != 'h3':
                    continue
                # Compruebo que el texto esté en negrita
                first_element = first_element.contents[0]
                if first_element.name != 'strong':
                    continue
                # Compruebo que el texto sea el correcto
                if first_element.string != 'Informe diario de situación':
                    continue

                # Es el elemento correcto, así que salgo del bucle
                break
            # Si no he conseguido acceder a alguno de los elementos,
            # es que no es la sección que me interesa.
            except:
                pass

        try:
            # Accedo al hipervínculo
            link_section = div.find('h2', {'class': "rtecenter"})
            # Tomo el link, texto y dirección
            link = link_section.find('a')
            # Extraigo la dirección
            sz_link = link.attrs['href']
        except:
            # Si no he podido acceder a alguno de los elementos,
            # no puedo descargar la página por tanto
            return None

        # Descargo la página
        response = requests.get(self.SZ_CAM_URL + sz_link)
        # Comrpuebo que la desarga haya ido bien
        if response.status_code == 200:
            # En este punto ya sé seguro que voy a devolver una página correcta.
            # Imprimo por pantalla el texto para decir qué fecha voy a extraer
            print(link.contents[0])
            return response
        else:
            return None

    def __try_combinations(self):
        # Defino la lista de prefijos
        self.__list_of_prefix()

        # Estoy intentando acceder al informe de hoy
        while True:
            # Pido el informe de hoy
            response = self.__get_date_response()
            if response.status_code == 200:
                # Aviso por pantalla de cuál es el último informe
                print("Último informe del día " + str(self.date.day) +
                  "-" + str(self.date.month) + "-" + str(self.date.year))
                # Devuelvo el contenido de la página a la que he conseguido acceder
                return response

            # Ninguna de las modulaciones es válida.
            # Asumo que aún no se ha publicado el informa diario.
            self.date = self.date - datetime.timedelta(days=1)

    def __list_of_prefix(self):
        # Lista de todas las variantes de prefijo hasta ahora
        prefix_list = [self.SZ_CAM_FILES + "/doc/sanidad/",
                       self.SZ_CAM_FILES + "/doc/sanidad/prev/",
                       self.SZ_CAM_FILES + "/aud/sanidad/prev/",
                       self.SZ_CAM_FILES + "/aud/sanidad/",
                       self.SZ_CAM_FILES + "/doc/presidencia/"]
        # Lista de todas las variantes de sufijo hasta ahora
        sufix_list = ["_cam_covid19",
                      "_cam_covid"]
        # Lista de todas las extensiones hasta ahora
        extension_list = [".pdf", ".pdf.pdf"]
        # Lista de cómo se ha formateado la fecha
        date_format_list = [lambda x: str(x.year)[-2:] + "{:02d}".format(x.month) + "{:02d}".format(x.day),
                            lambda x: str(x.year) + "{:02d}".format(x.month) + "{:02d}".format(x.day)]

        # Creo una lista donde guardo todas sus posibles combinaciones
        self.pre_sufix_list = []
        # Iteraciones ordenadas de más cambiante a más estable
        for extension in extension_list:
            for sufix in sufix_list:
                for prefix in prefix_list:
                    for date in date_format_list:
                        self.pre_sufix_list.append(URLFormat(prefix,
                                                             sufix,
                                                             extension,
                                                             date))
        return

    def __get_date_response(self):

        # Obtengo todas las variantes que puedo para el link de hoy
        links = [i.get_url(self.date) for i in self.pre_sufix_list]

        # De forma paralelizada descargo el contenido de todos los links
        with concurrent.futures.ThreadPoolExecutor() as executor:
            responses = list(executor.map(requests.get, links))

        # Miro cuántos de ellos me han devuelto realmente una página.
        # Me espero que sólo uno de ellos tengo response_code 200
        valid_resposes = [respose for respose in responses if respose.status_code == 200]
        if len(valid_resposes):
            # Devuelvo una página válida
            return valid_resposes[0]
        else:
            # Devuelvo una página no válida
            return responses[0]


class PDF_Reader():
    def __init__(self):
        # Descargo el último pdf disponible
        download = Downloader()
        download.download_pdf()
        self.pdf_name = download.pdf_name

        # Abro el archivo pdf en modo lectura
        self.pdf_file = open(self.pdf_name, 'rb')
        self.fileText = ""
        self.data = []
        return

    def __del__(self):
        # Cierro el archivo PDF
        self.pdf_file.close()
        # Elimino el PDF
        os.remove(self.pdf_name)

    def read_file(self):

        print("Abriendo PDF…")

        # creating a pdf reader object
        fileReader = pdfplumber.open(self.pdf_file)

        # Bool que me indica si he encontrado alguna tabla
        table_found = False

        print("Leyendo PDF…")
        for page in fileReader.pages:
            # Extraigo el texto de la página actual
            self.fileText = page.extract_text().replace('\n', '')

            # Compruebo si la página actual tiene tablas.
            if self.__has_tables():
                # Estoy en una página con tablas, leo su contenido.
                self.__get_clear_data()
                table_found = True
            # No estoy en una página con tablas.
            elif table_found:
                # Ya he leído todas las tablas.
                # Podemos dejar de leer.
                break

        # Ordeno los datos por fecha
        print("Ordenando datos…")
        self.data.sort(key=lambda tup: tup[0])

        return self.data

    def __has_tables(self):
        # Leo el encabezado de la página.
        # Compruebo si es el encabezado de una página con tablas.
        return self.fileText.find("Se realiza una actualización diaria ") >= 0

    def __get_clear_data(self):
        # Elimino la cabecera del texto
        self.fileText = self.fileText.split("Diario  Acumulado")[-1]
        # Elimino la cola del texto
        # La "F" es el inicio del pie de página:
        # "Fuente: Dirección General de Salud Pública"
        self.fileText = self.fileText.split("F")[0]
        # Guardo en una lista todos los datos, no están ordenados
        data = self.fileText.split(" ")
        # Elimino las cadenas vacías
        data = [i for i in data if i]
        # Por si ha entrado algún texto que no quería
        data = self.__check_header(data)
        # Obtengo los pares eliminando los agregados
        data = list(zip(data[::3], data[1::3]))
        # Convierto los datos a fecha y enteros
        data = [[datetime.datetime.strptime(
            i[0], "%d/%m/%Y"), int(i[1])] for i in data]
        # Actualizo la variable miembro con lo leído en la página actual
        self.data = self.data + data

    def __check_header(self, list_page):
        # Los encabezados de la tabla siempre estarán al principio de mi lista.
        # Leo su primer elemento hasta que se afectivamente una fecha.
        while list_page:
            try:
                # Compruebo que el primer dato sea una fecha…
                datetime.datetime.strptime(list_page[0], "%d/%m/%Y")
            except:
                # …en caso contrario elimino este elemento
                list_page.pop(0)
            else:
                # Si en efecto es una fecha, no hay nada más que modificar en el encabezado.
                break

        return list_page


class Excel_writer():
    def __init__(self):
        self.excel_name = "Casos Comunidad de Madrid.xlsx"
        print("Abriendo Excel…")
        self.workbook = load_workbook(self.excel_name)

        # Guardo la hoja donde voy a escribir los datos.
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # Las celdas se numeran a partir de 1.
        # No puedo escribir en la primera fila, es para los encabezados.
        self.index_line = 2

    def __del__(self):
        print("Guardando Excel…")
        # Guardo el xlsx.
        # Le doy el mismo nombre que tenía.
        self.workbook.save(self.excel_name)
        # Cierro el xlsx
        self.workbook.close()

    def write_data(self, data):
        # Data debe ser una lista de pares.
        # El primer elemento debe ser una fecha en formato fecha.
        # El segundo elemento, la cantidad de positivos en formato entero.
        print("Escribiendo Excel…")
        for value in data:
            # Fecha escrita con formato de fecha.
            self.__write_date(value)
            # Positivos escritos con formato de entero.
            self.__write_cases(value)

            # Avanzo a la siguiente linea.
            self.index_line = self.index_line + 1

    def __write_date(self, value):
        # Dado el par value, escribo la fecha (1º elemento) en la casilla adecuada de excel.

        cell = self.sheet.cell(row=self.index_line, column=1)
        # Fecha escrita con formato de fecha.
        cell.value = value[0]
        cell.number_format = 'mm-dd-yy'

    def __write_cases(self, value):
        # Dado el par value, escribo los positivos (2º elemento) en la casilla adecuada de excel.

        cell = self.sheet.cell(row=self.index_line, column=2)
        # Positivos escritos con formato de entero.
        cell.value = value[1]
        cell.number_format = '#,##0'


if __name__ == "__main__":

    last_data = PDF_Reader()
    list_data = last_data.read_file()
    del last_data

    excel = Excel_writer()
    excel.write_data(list_data)
    del excel
