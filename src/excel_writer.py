import openpyxl
import enum


class ExcelColumns(enum.Enum):
    Fecha = enum.auto()
    Positivos = enum.auto()
    Media_7_dias = enum.auto()
    Derivada_7 = enum.auto()
    Media_14_dias = enum.auto()
    Derivada_14 = enum.auto()
    Promedio = enum.auto()
    Linea_records = enum.auto()
    Riesgo_bajo = enum.auto()
    Riesgo_medio = enum.auto()
    Riesgo_alto = enum.auto()
    Riesgo_extremo = enum.auto()
    Media_derivada = enum.auto()
    Media_reproductivo = enum.auto()
    Dia_semana = enum.auto()
    Suma_parcial = enum.auto()
    Reproductivo = enum.auto()

    def __str__(self) -> str:
        return f"Tabla2[[#This Row],[{self.name}]]"

    def __int__(self) -> int:
        return self.value


class Excel_writer():
    def __init__(self):
        self.excel_name = "Casos Comunidad de Madrid.xlsx"
        print("Abriendo Excel…")
        self.workbook = openpyxl.load_workbook(self.excel_name)

        # Guardo la hoja donde voy a escribir los datos.
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # Las celdas se numeran a partir de 1.
        # No puedo escribir en la primera fila, es para los encabezados.
        self.index_line = 2

    def __del__(self):
        print("Guardando Excel…")
        # Guardo el xlsx.
        # Le doy el mismo nombre que tenía.
        self.workbook.save(self.excel_name)
        # Cierro el xlsx
        self.workbook.close()

    def write_data(self, data):
        # Data debe ser una lista de pares.
        # El primer elemento debe ser una fecha en formato fecha.
        # El segundo elemento, la cantidad de positivos en formato entero.
        print("Escribiendo Excel…")
        for value in data:
            # Fecha escrita con formato de fecha.
            self.__write_date(value)
            # Positivos escritos con formato de entero.
            self.__write_cases(value)
            # Resto de columnas
            self.__write_other_columns()

            # Avanzo a la siguiente linea.
            self.index_line = self.index_line + 1

    def __write_date(self, value):
        # Dado el par value, escribo la fecha (1º elemento) en la casilla adecuada de excel.

        cell = self.sheet.cell(row=self.index_line,
                               column=int(ExcelColumns.Fecha))
        # Fecha escrita con formato de fecha.
        cell.value = value[0]
        cell.number_format = 'mm-dd-yy'

    def __write_cases(self, value):
        # Dado el par value, escribo los positivos (2º elemento) en la casilla adecuada de excel.

        cell = self.sheet.cell(row=self.index_line,
                               column=int(ExcelColumns.Positivos))
        # Positivos escritos con formato de entero.
        cell.value = value[1]
        cell.number_format = '#,##0'

    def __write_other_columns(self):
        # Escribo la media de últimos 7 y 14 días
        formula = f"=AVERAGE(OFFSET({ExcelColumns.Positivos},0,0,-MIN(ROW()-1,7)))"
        self.__set_cell_value(ExcelColumns.Media_7_dias,
                              formula)
        formula = f"=AVERAGE(OFFSET({ExcelColumns.Positivos},0,0,-MIN(ROW()-1,14)))"
        self.__set_cell_value(ExcelColumns.Media_14_dias,
                              formula)

        # Escribo la suma parcial
        formula = f"=SUM(OFFSET({ExcelColumns.Positivos}, 0, 0, -ROW()))"
        self.__set_cell_value(ExcelColumns.Suma_parcial,
                              formula)

        # Escribo el día de la semana
        formula = f"=TEXT({ExcelColumns.Fecha},\"dddd\")"
        self.__set_cell_value(ExcelColumns.Dia_semana,
                              formula)

        # Escribo el número reproductivo
        if self.index_line == 2:
            formula = "=1"
        else:
            formula = f"={ExcelColumns.Media_14_dias}/OFFSET({ExcelColumns.Media_14_dias},-1,0)"
        self.__set_cell_value(ExcelColumns.Reproductivo,
                              formula)
        # Escribo las derivadas
        if self.index_line != 2:
            formula = f"={ExcelColumns.Media_14_dias}-OFFSET({ExcelColumns.Media_14_dias},-1,0)"
        self.__set_cell_value(ExcelColumns.Derivada_14,
                              formula)
        if self.index_line != 2:
            formula = f"={ExcelColumns.Media_7_dias}-OFFSET({ExcelColumns.Media_7_dias},-1,0)"
        self.__set_cell_value(ExcelColumns.Derivada_7,
                              formula)
        # Media de la derivada
        formula = f"=AVERAGE(OFFSET({ExcelColumns.Derivada_7}, 0, 0, -MIN(ROW() - 1,7)))"
        self.__set_cell_value(ExcelColumns.Media_derivada,
                              formula)
        # Media del número reproductivo
        formula = f"=GEOMEAN(OFFSET({ExcelColumns.Reproductivo}, 0, 0, -MIN(ROW() - 1,7)))"
        self.__set_cell_value(ExcelColumns.Media_reproductivo,
                              formula)

        # Escribo la línea de récord
        formula = f"=IF({ExcelColumns.Fecha}<m_FechaObjetivo,NA(),m_UltimoDato14Dias)"
        self.__set_cell_value(ExcelColumns.Linea_records,
                              formula)

        # Escribo las líneas de riesgo
        self.__set_cell_value(ExcelColumns.Riesgo_bajo,
                              "=m_RiesgoBajo")
        self.__set_cell_value(ExcelColumns.Riesgo_medio,
                              "=m_RiesgoMedio")
        self.__set_cell_value(ExcelColumns.Riesgo_alto,
                              "=m_RiesgoAlto")
        self.__set_cell_value(ExcelColumns.Riesgo_extremo,
                              "=m_RiesgoExtremo")
        self.__set_cell_value(ExcelColumns.Promedio,
                              f"=AVERAGE([{ExcelColumns.Positivos.name}])")

    def __set_cell_value(self, col: ExcelColumns, value: str):

        cell = self.sheet.cell(row=self.index_line,
                               column=int(col))
        cell.value = value

        if (col == ExcelColumns.Media_7_dias or
            col == ExcelColumns.Derivada_7 or
            col == ExcelColumns.Media_14_dias or
                col == ExcelColumns.Derivada_14):
            cell.number_format = '#,##0'

        elif (col == ExcelColumns.Dia_semana):
            # Formato de texto
            pass

        elif (col == ExcelColumns.Reproductivo):
            cell.number_format = '0.00'

        elif (col == ExcelColumns.Riesgo_bajo or
              col == ExcelColumns.Riesgo_medio or
              col == ExcelColumns.Riesgo_alto or
              col == ExcelColumns.Riesgo_extremo or
              col == ExcelColumns.Promedio):
            cell.number_format = '0.00E+00'
